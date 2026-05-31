"""
Generuje raport XLSX z wynikami walidacji rozmytego systemu diagnostyki cukrzycy.
Wejscie:  results_20.csv  (generowany przez DiabetesTest.java)
Wyjscie:  Analiza_20_przypadkow.xlsx
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Kolory ──────────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # ciemny niebieski
C_SUBHDR   = "2E75B6"   # sredni niebieski
C_OK       = "E2EFDA"   # jasna zielen
C_BLAD     = "FCE4D6"   # jasny rozowy
C_NISKIE   = "E2EFDA"
C_UMIAR    = "FFEB9C"
C_WYSOK    = "F4B942"
C_BWYSOK   = "FF0000"
C_ALT1     = "DDEEFF"
C_ALT2     = "FFFFFF"

def hdr_font(bold=True, color="FFFFFF", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def risk_fill(level):
    m = {"NISKIE": C_NISKIE, "UMIARKOWANE": C_UMIAR,
         "WYSOKIE": C_WYSOK, "BARDZO WYSOKIE": C_BWYSOK}
    return fill(m.get(level, "FFFFFF"))

def risk_font(level):
    color = "FFFFFF" if level == "BARDZO WYSOKIE" else "000000"
    return cell_font(bold=True, color=color)

# ── Wczytaj CSV ──────────────────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
csv_path   = os.path.join(script_dir, "bin", "results_20.csv")
xlsx_path  = os.path.join(script_dir, "Analiza_20_przypadkow.xlsx")

rows = []
with open(csv_path, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# ARKUSZ 1: Tabela wynikow
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Wyniki 20 przypadkow"
ws1.freeze_panes = "A3"

# -- Tytul --
ws1.merge_cells("A1:R1")
title_cell = ws1["A1"]
title_cell.value = "Walidacja rozmytego systemu diagnostyki cukrzycy ciazowej – pierwsze 20 przypadkow (Pima Indians Diabetes Dataset)"
title_cell.font  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
title_cell.fill  = fill(C_HEADER)
title_cell.alignment = center()
ws1.row_dimensions[1].height = 28

# -- Naglowki --
headers = [
    "Nr", "Ciaze\n(l. ciaż)", "Glukoza\n(mg/dl)", "Cisnienie\n(mmHg)",
    "Grubosc skory\n(mm)", "BMI", "Wiek\n(lat)",
    "Wynik\nrzeczywisty", "Ryzyko\n[%]", "Poziom\nryzyka",
    "Predykcja\nsystemu", "Poprawny?",
    "Glukoza\n(etyk.)", "BMI\n(etyk.)", "Cisnienie\n(etyk.)",
    "L. ciaż\n(etyk.)", "Wiek\n(etyk.)", "Skora\n(etyk.)"
]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font      = hdr_font()
    c.fill      = fill(C_SUBHDR)
    c.alignment = center()
    c.border    = thin_border()
ws1.row_dimensions[2].height = 36

# -- Dane --
correct = 0
for i, r in enumerate(rows):
    row = i + 3
    bg = C_ALT1 if i % 2 == 0 else C_ALT2
    ok = r["Poprawny"] == "TAK"
    if ok:
        correct += 1

    values = [
        int(r["Nr"]),
        float(r["Ciaze"]),
        float(r["Glukoza"]),
        float(r["Cisnienie"]),
        float(r["Skora"]),
        float(r["BMI"]),
        float(r["Wiek"]),
        "CHORA" if r["Outcome"] == "1" else "ZDROWA",
        round(float(r["Ryzyko"]), 1),
        r["PoziomRyzyka"],
        "CHORA" if r["Predykcja"] == "1" else "ZDROWA",
        "TAK ✓" if ok else "NIE ✗",
        r["EtykietaGl"],
        r["EtykietaBMI"],
        r["EtykietaCi"],
        r["EtykietaCiaz"],
        r["EtykietaWiek"],
        r["EtykietaSkora"],
    ]
    for col, val in enumerate(values, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.border    = thin_border()
        c.alignment = center()

        # Kolorowanie specjalne
        if col == 9:   # Ryzyko %
            c.font = cell_font()
            c.fill = risk_fill(r["PoziomRyzyka"])
            c.font = risk_font(r["PoziomRyzyka"])
        elif col == 10:  # Poziom ryzyka
            c.fill = risk_fill(r["PoziomRyzyka"])
            c.font = risk_font(r["PoziomRyzyka"])
        elif col == 12:  # Poprawny
            c.fill = fill(C_OK) if ok else fill(C_BLAD)
            c.font = cell_font(bold=True, color="276221" if ok else "9C0006")
        elif col in (8, 11):  # Wynik rzeczywisty / predykcja
            is_chora = val == "CHORA"
            c.fill = fill("FCE4D6") if is_chora else fill("E2EFDA")
            c.font = cell_font(bold=True)
        else:
            c.fill = fill(bg)
            c.font = cell_font()

    ws1.row_dimensions[row].height = 18

# -- Wiersz podsumowania --
sum_row = len(rows) + 3
ws1.merge_cells(f"A{sum_row}:G{sum_row}")
ws1.cell(sum_row, 1).value = "PODSUMOWANIE"
ws1.cell(sum_row, 1).font  = hdr_font(size=10)
ws1.cell(sum_row, 1).fill  = fill(C_HEADER)
ws1.cell(sum_row, 1).alignment = center()

acc_pct = correct * 5.0
ws1.merge_cells(f"H{sum_row}:L{sum_row}")
ws1.cell(sum_row, 8).value = f"Poprawnych: {correct}/20  ({acc_pct:.0f}%)"
ws1.cell(sum_row, 8).font  = hdr_font(size=11)
ws1.cell(sum_row, 8).fill  = fill(C_SUBHDR)
ws1.cell(sum_row, 8).alignment = center()
ws1.row_dimensions[sum_row].height = 22

# -- Szerokosci kolumn --
col_widths = [5, 10, 11, 12, 14, 8, 8, 12, 10, 16, 14, 11,
              14, 13, 14, 13, 13, 15]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# ARKUSZ 2: Analiza statystyczna
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Analiza statystyczna")

def section(ws, row, text, cols=2):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = hdr_font(size=11)
    c.fill = fill(C_SUBHDR)
    c.alignment = center()
    c.border = thin_border()
    ws.row_dimensions[row].height = 22

def stat_row(ws, row, label, value, bg=C_ALT2):
    c1 = ws.cell(row, 1, label)
    c2 = ws.cell(row, 2, value)
    for c in (c1, c2):
        c.border    = thin_border()
        c.alignment = left()
        c.fill      = fill(bg)
        c.font      = cell_font()
    c2.alignment = center()
    ws.row_dimensions[row].height = 18

# Tytuł
ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value = "Analiza statystyczna – system rozmyty diagnostyki cukrzycy ciazowej"
c.font  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
c.fill  = fill(C_HEADER)
c.alignment = center()
ws2.row_dimensions[1].height = 28

# Zlicz
tp=tn=fp=fn=0
risks_0, risks_1 = [], []
risk_levels = {"NISKIE":0,"UMIARKOWANE":0,"WYSOKIE":0,"BARDZO WYSOKIE":0}
for r in rows:
    risk = float(r["Ryzyko"])
    pred = int(r["Predykcja"])
    actual = int(r["Outcome"])
    level = r["PoziomRyzyka"]
    risk_levels[level] = risk_levels.get(level, 0) + 1
    if actual == 0:
        risks_0.append(risk)
    else:
        risks_1.append(risk)
    if pred==1 and actual==1: tp+=1
    elif pred==0 and actual==0: tn+=1
    elif pred==1 and actual==0: fp+=1
    else: fn+=1

acc  = 100.0*(tp+tn)/20
prec = 100.0*tp/(tp+fp) if (tp+fp)>0 else 0
rec  = 100.0*tp/(tp+fn) if (tp+fn)>0 else 0
f1   = 2*prec*rec/(prec+rec) if (prec+rec)>0 else 0
avg0 = sum(risks_0)/len(risks_0) if risks_0 else 0
avg1 = sum(risks_1)/len(risks_1) if risks_1 else 0

row = 3
section(ws2, row, "Macierz pomyłek", 4); row+=1

ws2.merge_cells(f"B{row}:B{row}")
ws2.merge_cells(f"C{row}:D{row}")
for col, txt in [(2,"Pred: ZDROWA"),(3,"Pred: CHORA")]:
    c = ws2.cell(row, col, txt)
    c.font = hdr_font(size=10); c.fill = fill(C_SUBHDR)
    c.alignment = center(); c.border = thin_border()
ws2.cell(row, 1, "").fill = fill(C_SUBHDR)
row+=1

for label, left_val, right_val, left_ok in [
        ("Rzeczyw: ZDROWA", f"TN = {tn}", f"FP = {fp}", True),
        ("Rzeczyw: CHORA",  f"FN = {fn}", f"TP = {tp}", False)]:
    c1 = ws2.cell(row, 1, label); c1.font=hdr_font(10,C_HEADER)
    c1.fill=fill(C_ALT1); c1.alignment=center(); c1.border=thin_border()
    c2 = ws2.cell(row, 2, left_val)
    c3 = ws2.cell(row, 3, right_val)
    c2.fill = fill(C_OK if left_ok else C_BLAD)
    c3.fill = fill(C_BLAD if left_ok else C_OK)
    for c in (c2, c3):
        c.font=cell_font(bold=True); c.alignment=center(); c.border=thin_border()
    ws2.merge_cells(f"C{row}:D{row}")
    row+=1

row+=1
section(ws2, row, "Metryki klasyfikacji", 2); row+=1
for lbl, val, bg in [
    ("Próg decyzyjny",     "45%",              C_ALT1),
    ("Dokładność (Accuracy)",  f"{acc:.1f}%",  C_ALT2),
    ("Precyzja (Precision)",   f"{prec:.1f}%", C_ALT1),
    ("Czułość (Recall/SE)",    f"{rec:.1f}%",  C_ALT2),
    ("F1-score",               f"{f1:.1f}%",   C_ALT1),
    ("Śr. ryzyko – zdrowe (Outcome=0)", f"{avg0:.1f}%  (n={len(risks_0)})", C_ALT2),
    ("Śr. ryzyko – chore  (Outcome=1)", f"{avg1:.1f}%  (n={len(risks_1)})", C_ALT1),
]:
    stat_row(ws2, row, lbl, val, bg); row+=1

row+=1
section(ws2, row, "Rozkład poziomów ryzyka (20 przypadków)", 2); row+=1
for level, cnt in risk_levels.items():
    c1 = ws2.cell(row, 1, level)
    c2 = ws2.cell(row, 2, cnt)
    c1.fill = risk_fill(level); c1.font = risk_font(level)
    c2.fill = risk_fill(level); c2.font = risk_font(level)
    for c in (c1,c2):
        c.border=thin_border(); c.alignment=center()
    ws2.row_dimensions[row].height = 18
    row+=1

row+=1
section(ws2, row, "Obserwacje i wnioski", 3); row+=1
notes = [
    "Brakujące dane (zero) – dataset Pima używa 0 jako placeholder dla brakujących wartości cisnienia, grubości skóry i BMI.",
    "Para1 dead zones – reguły wymagają zgodności poziomów gl i BMI; wysoka glukoza z normalnym BMI = Para1≈0 → wynik 50%.",
    "Wynik 50.0% – COG przy braku dominującej reguły; to artefakt biblioteki fuzzlib gdy żadna reguła nie odpali mocno.",
    "Próg decyzyjny 45% daje 65% celności; wyższy próg poprawia czułość kosztem precyzji.",
    "System działa poprawnie dla 'typowych' przypadków – zgodność profilu gl+BMI z regułami.",
]
for note in notes:
    ws2.merge_cells(f"A{row}:D{row}")
    c = ws2.cell(row, 1, f"• {note}")
    c.font = cell_font(size=9); c.fill = fill(C_ALT1)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws2.row_dimensions[row].height = 40
    row+=1

for col, w in [(1,32),(2,14),(3,14),(4,14)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════════
# ARKUSZ 3: Wykres słupkowy
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Wykres ryzyka")

ws3.merge_cells("A1:J1")
c = ws3["A1"]
c.value = "Ryzyko cukrzycy rozmytego systemu vs. wynik rzeczywisty – 20 przypadków"
c.font  = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
c.fill  = fill(C_HEADER)
c.alignment = center()
ws3.row_dimensions[1].height = 26

headers3 = ["Nr", "Ryzyko [%]", "Wynik rzecz. (0/1)"]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(2, col, h)
    c.font=hdr_font(size=10); c.fill=fill(C_SUBHDR)
    c.alignment=center(); c.border=thin_border()
ws3.row_dimensions[2].height = 20

for i, r in enumerate(rows):
    row = i+3
    nr   = int(r["Nr"])
    risk = round(float(r["Ryzyko"]),1)
    out  = int(r["Outcome"])
    for col, val in [(1,nr),(2,risk),(3,out*100)]:
        c = ws3.cell(row, col, val)
        c.alignment = center(); c.border = thin_border()
        c.fill = fill(C_ALT1 if i%2==0 else C_ALT2)
        c.font = cell_font()
    ws3.row_dimensions[row].height = 16

# Wykres
chart = BarChart()
chart.type   = "col"
chart.title  = "Ryzyko systemu vs Wynik rzeczywisty"
chart.y_axis.title = "Wartość [%]"
chart.x_axis.title = "Nr przypadku"
chart.style  = 10
chart.width  = 20
chart.height = 12

data_risk = Reference(ws3, min_col=2, min_row=2, max_row=22)
data_out  = Reference(ws3, min_col=3, min_row=2, max_row=22)
cats      = Reference(ws3, min_col=1, min_row=3, max_row=22)

chart.add_data(data_risk, titles_from_data=True)
chart.add_data(data_out,  titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = "2E75B6"
chart.series[1].graphicalProperties.solidFill = "FF6B6B"

ws3.add_chart(chart, "E3")

for col, w in [(1,8),(2,14),(3,18)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

# ── Zapis ────────────────────────────────────────────────────────────────────
wb.save(xlsx_path)
print(f"Zapisano: {xlsx_path}")
print(f"Poprawnych: {correct}/20 ({correct*5}%)  |  TP={tp} TN={tn} FP={fp} FN={fn}")
print(f"Accuracy={acc:.1f}%  Precision={prec:.1f}%  Recall={rec:.1f}%  F1={f1:.1f}%")
